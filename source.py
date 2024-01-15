from openpyxl.styles import Alignment, Font, PatternFill, NamedStyle
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from time import sleep
from openpyxl import Workbook

from listingfetcher import fetchListingsAndPopulateWorksheet
from averagepricefetcher import fetchAveragePrices
from helper import saveWorkbook, calculatePriceDifference, getListingPriceDifferenceColor

service = Service()
option = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=option)

listingUrl = "https://www.ss.lv/lv/electronics/computers/completing-pc/"
averagePricesUrl = "https://bestvaluegpu.com/"

driver.get(listingUrl)
sleep(1)

# Create new Excel workbook for output
wb = Workbook()
ws = wb.active

# Set Excel header field values
ws.title = 'Videokartes'
ws['A1'] = 'Modelis'
ws['B1'] = 'Zīmols'
ws['C1'] = 'Lietota'
ws['D1'] = 'Cena'
ws['E1'] = 'Attēls'
ws['F1'] = 'Saite'

# Fetch listings
fetchListingsAndPopulateWorksheet(driver, ws)

# Fetch average prices
driver.get(averagePricesUrl)

newAveragePrices = fetchAveragePrices(driver)

# Switch to used graphic card average prices
driver.find_element(By.XPATH, "//button[text()='Condition:']").click()
driver.find_element(By.XPATH, "//ul[@aria-label='Table Columns']/li[2]").send_keys(Keys.ENTER)
driver.find_element(By.XPATH, "//ul[@aria-label='Table Columns']/li[1]").send_keys(Keys.ENTER)
driver.find_element(By.XPATH, "//*").click()

usedAveragePrices = fetchAveragePrices(driver)

FormatEuro = NamedStyle(name='euro_style', number_format='€#,##0.00')

# Comparing listing prices with average prices
for row in ws.iter_rows(min_row=2):
    listingModel = row[0].value.lower()
    listingUsed = row[2] == "Jā"
    listingPrice = row[3].value
    listingPriceDifference = calculatePriceDifference(usedAveragePrices, listingModel, listingPrice) if row[2] == "Jā" else calculatePriceDifference(newAveragePrices, listingModel, listingPrice)

    # Set price color based on the difference with average price for this video card
    listingPriceColor = getListingPriceDifferenceColor(listingPriceDifference)
    row[3].style = FormatEuro
    row[3].fill = PatternFill(start_color=listingPriceColor, end_color=listingPriceColor, fill_type='solid')

    # Height and merge cell formatting
    ws.row_dimensions[row[0].row].height = 50

# Excel file graphical formatting

# Header formatting
for cell in ws[1]:
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(bold=True, size=18)

# Width formatting
for col in ws.columns:
    maxLength = 0
    letter = col[0].column_letter
    for cell in col:
        value = cell.value
        if value is not None and len(str(value)) > maxLength:
            maxLength = len(value)
        updatedWidth = (maxLength + 2) * 1.4
        ws.column_dimensions[letter].width = updatedWidth

# Save the output workbook
saveWorkbook(wb)