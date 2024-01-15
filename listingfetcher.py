from selenium.webdriver.common.by import By
from time import sleep
from openpyxl.worksheet.hyperlink import Hyperlink
from io import BytesIO
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, NamedStyle

from helper import acceptCookiesIfPromptPresent, containsValue

categoryXPath = "//span[@class='filter_opt_dv'][4]/select"
dealTypeXPath = "//span[@class='filter_opt_dv'][3]/select"
listingXPath = "//div[@class='filter_second_line_dv']/following::table[1]//tr[position()>1 and position()<last()]"
includedModels = [
    "rtx", "rx"
]
excludedModels = [
    "6500", "6600", "5700", "5600", "3050"
]
excludedRegions = [
    "Liepāja un raj.", "Daugavpils un raj.", "Rēzekne un raj.", "Bauska un raj."
]

def fetchListingsAndPopulateWorksheet(driver, ws):
    # Listings start from second entry since first is the header
    listingRow = 2
    hyperlinkStyle = NamedStyle(name='hyperlink', font=Font(underline='single', color='0563C1'))

    # Select category
    driver.find_element(By.XPATH, categoryXPath).click()
    driver.find_element(By.XPATH, categoryXPath + "/option[text()='Video']").click()

    # Fetch page count
    pageCount = int(driver.find_element(By.XPATH, "//a[@class='navi'][last() - 1]").text)

    # Not using enumerate with listingRow because represented listing count in Excel will not be always equal to the fetched listing count
    for i in range(pageCount):

        # Show only "Sell" listings
        driver.find_element(By.XPATH, dealTypeXPath).click()
        driver.find_element(By.XPATH, dealTypeXPath + "/option[text()='Pārdod']").click()

        # Fetch listings
        listings = driver.find_elements(By.XPATH, listingXPath)

        for listing in listings:
            acceptCookiesIfPromptPresent(driver)
            listingUrl = listing.find_element(By.XPATH, ".//td[2]/a").get_attribute("href")
            listingPictureData = listing.find_element(By.XPATH, ".//td[2]/a/img").screenshot_as_png
            listingRegion = listing.find_element(By.XPATH, ".//td[3]/div[@class='ads_region']").text
            listingBrand, listingModel = listing.find_element(By.XPATH, ".//td[4]").text.split("\n")
            listingUsed = "Jā" if listing.find_element(By.XPATH, ".//td[7]").text == "lietota" else "Nē"
            listingPrice = listing.find_element(By.XPATH, ".//td[8]").text
            listingPrice = int(listingPrice[:listingPrice.find(' €')].replace(',', ''))
            if containsValue(listingModel, includedModels) and not containsValue(listingModel, excludedModels) and not containsValue(listingRegion, excludedRegions) and 100 < listingPrice < 1000:
                ws[f'A{listingRow}'] = listingModel
                ws[f'B{listingRow}'] = listingBrand
                ws[f'C{listingRow}'] = listingUsed
                ws[f'D{listingRow}'] = listingPrice
                listingPictureStream = BytesIO(listingPictureData)
                listingPicture = Image(listingPictureStream)
                ws.add_image(listingPicture, f'E{listingRow}')
                ws[f'F{listingRow}'] = listingUrl
                ws[f'F{listingRow}'].style = hyperlinkStyle
                ws[f'F{listingRow}'].hyperlink = Hyperlink(ref=listingUrl, target=listingUrl)
                listingRow += 1

        # Navigate to next page
        driver.find_element(By.XPATH, "//a[@class='navi'][last()]").click()
        sleep(0.25)